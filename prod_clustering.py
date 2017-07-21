# -*- coding: utf-8 -*-
import openpyxl
import os.path
import numpy as np
import sys
from sklearn import preprocessing
from sklearn.cluster import KMeans
from sklearn import datasets
import matplotlib.pyplot as plt
from sklearn.externals import joblib
from sklearn.cluster import AffinityPropagation
from sklearn import metrics
from sklearn.datasets.samples_generator import make_blobs
import random
from sklearn.cluster import MeanShift, estimate_bandwidth
from itertools import cycle
from sklearn.metrics import silhouette_samples, silhouette_score

def kmeans_clustering(np_attr,output_to_excel,code,elbow ):
    s = []
    if elbow == True:
        Ks = range(1,20)
        km = [KMeans(n_clusters=i) for i in Ks]
        score = [km[i].fit(np_attr).score(np_attr) for i in range(len(km))]
        plt.plot(Ks, score)
        plt.show()
    else:
        true_k = 4
        model = KMeans(n_clusters=true_k, init="k-means++", random_state=170, max_iter=100, n_init=1)
        model.fit(np_attr)

        clusters = model.labels_.tolist()
        order_centroids = model.cluster_centers_.argsort()[:,::-1]
        print("The cluster centroids are: \n", model.cluster_centers_)
        print("Cluster", model.labels_)
        clusters = model.labels_
        print()
        print("Sum of distances of samples to their closest cluster center: ", model.inertia_)
        if output_to_excel == True:
            output(model.labels_,code)


def normalizing(np_attr):
    np_attr = np_attr.astype(np.float)
    np_attr = preprocessing.scale(np_attr)
    return np_attr

def affinity_propagation_clustering(np_attr,keys,normalized,code,output_to_excel):
    # normalize data, see http://scikit-learn.org/stable/modules/preprocessing.html for differet methods
    if normalized == True:
        np_attr = normalizing(np_attr)
    #centers = [[2, 2], [8, 9], [9, 5], [3,9],[4,4],[0,0],[2,5]]
    #x, labels_true = make_blobs(n_samples=1000, centers=centers, cluster_std=0.5,random_state=0)

    #labels_true = random.sample(range(1, 1059), 1058)
    labels_true = keys
    af = AffinityPropagation(preference=-100).fit(np_attr)
    cluster_centers_indices = af.cluster_centers_indices_
    labels = af.labels_
    print(len(set(labels)))
    n_clusters_ = len(cluster_centers_indices)
    print('Estimated number of clusters: %d' % n_clusters_)
    print("Homogeneity: %0.3f" % metrics.homogeneity_score(labels_true, labels))
    print("Completeness: %0.3f" % metrics.completeness_score(labels_true, labels))
    print("V-measure: %0.3f" % metrics.v_measure_score(labels_true, labels))
    print("Adjusted Rand Index: %0.3f"
          % metrics.adjusted_rand_score(labels_true, labels))
    print("Adjusted Mutual Information: %0.3f"
          % metrics.adjusted_mutual_info_score(labels_true, labels))
    print("Silhouette Coefficient: %0.3f"
          % metrics.silhouette_score(np_attr, labels, metric='sqeuclidean'))
    # set colors for the clusters


    plt.close('all')
    plt.figure(1)
    plt.clf()

    colors = cycle('bgrcmykbgrcmykbgrcmykbgrcmyk')
    for k, col in zip(range(n_clusters_), colors):
        class_members = labels == k
        cluster_center = np_attr[cluster_centers_indices[k]]
        plt.plot(np_attr[class_members, 0], np_attr[class_members, 1], col + '.')
        plt.plot(cluster_center[0], cluster_center[1], 'o', markerfacecolor=col,
                 markeredgecolor='k', markersize=14)
        for x in np_attr[class_members]:
            plt.plot([cluster_center[0], x[0]], [cluster_center[1], x[1]], col)

    plt.title('Estimated number of clusters: %d' % n_clusters_)
    plt.show()
    if output_to_excel == True:
        output(af.labels_,code)

def meanshift_clustering(np_attr,keys,normalized,code,output_to_excel):
    if normalized == True:
        np_attr = normalizing(np_attr)
    bandwidth = estimate_bandwidth(np_attr, quantile=0.2, n_samples=1058)
    ms = MeanShift(bandwidth=bandwidth, bin_seeding=True)
    ms.fit(np_attr)
    labels = ms.labels_
    cluster_centers = ms.cluster_centers_

    labels_unique = np.unique(labels)
    n_clusters_ = len(labels_unique)

    print("number of estimated clusters : %d" % n_clusters_)
    plt.figure(1)
    plt.clf()

    colors = cycle('bgrcmykbgrcmykbgrcmykbgrcmyk')
    for k, col in zip(range(n_clusters_), colors):
        my_members = labels == k
        cluster_center = cluster_centers[k]
        plt.plot(np_attr[my_members, 0], np_attr[my_members, 1], col + '.')
        plt.plot(cluster_center[0], cluster_center[1], 'o', markerfacecolor=col,
                 markeredgecolor='k', markersize=14)
    plt.title('Estimated number of clusters: %d' % n_clusters_)
    plt.show()
    if output_to_excel == True:
        output(ms.labels_,code)

def output(labels,code):
    time = {}
    keys = list(code.keys())
    clusters = {}

    i = 0
    for name in code.keys():
        if labels[i] not in clusters.keys():
            clusters[labels[i]] = {}
            clusters[labels[i]]['products']= []
            clusters[labels[i]]['time'] = 0
            clusters[labels[i]]['products'].append(name)
            clusters[labels[i]]['time'] += code[name]['prod_time']

        else:
            clusters[labels[i]]['products'].append(keys[i])
            clusters[labels[i]]['time'] += code[name]['prod_time']
        i += 1

    if os.path.isfile("main.xlsx") == True:
        os.remove("main.xlsx")

    book = openpyxl.Workbook()
    sheet = book.active
    print(clusters.keys())
    for i in range(0, len(clusters)):
        sheet.cell(row=0, column=i).value = "cluster" + str(i)
        clusters[i]['time'] = clusters[i]['time']/len(clusters[i]['products'])
        for j in range(0, len(clusters[i]['products'])):
            sheet.cell(row=j + 1, column=i).value = clusters[i]['products'][j]
        sheet.cell(row=i, column=len(clusters) + 5).value = "cluster" + str(i)
        sheet.cell(row=i, column=len(clusters) + 6).value = clusters[i]['time']
    book.save("main.xlsx")

wb = openpyxl.load_workbook('production_data.xlsx')
data = wb.get_sheet_by_name("Sheet1")
length = 1063
i = 1
code = {}
attribute = []

while data.cell(row=i, column=0).value is not None:
    if data.cell(row=i, column=0).value not in code.keys():
        code[(data.cell(row=i, column=0).value)] = {}
        temp = []
        for k in range (2,6):
            if k == 2:
                code[(data.cell(row=i, column=0).value)]["batch_size"] = data.cell(row=i, column=k).value
                temp.append(data.cell(row=i, column=k).value)
            elif k == 3:
                code[(data.cell(row=i, column=0).value)]["prod_time"] = data.cell(row=i, column=k).value
                temp.append(data.cell(row=i, column=k).value)
            elif k == 4:
                code[(data.cell(row=i, column=0).value)]["danger_level"] = data.cell(row=i, column=k).value
                temp.append(data.cell(row=i, column=k).value)
            else:
                code[(data.cell(row=i, column=0).value)]["freq"] = data.cell(row=i, column=k).value
                temp.append(data.cell(row=i, column=k).value)

        #print(temp)
        attribute.append(temp)
    i += 1
#pprint.pprint(attribute)

keys = list(code.keys())
np_attr=np.asarray(attribute)
#meanshift_clustering(np_attr,keys,True,code,True)
#affinity_propagation_clustering(np_attr,keys,True,code,True)
kmeans_clustering(np_attr,False,code,True)


#plt.plot(initial)
#plt.show()

